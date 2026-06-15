import cv2
import easyocr
import numpy as np
import csv
import os
import time
import re
import openpyxl

BASE_DIR = "Solar_Cells"
EXCEL_FILE = "serial_numbers.xlsx"

os.makedirs(BASE_DIR, exist_ok=True)

reader = easyocr.Reader(['en']) # loads ocr model

cam_back = cv2.VideoCapture(1, cv2.CAP_DSHOW)   # back camera
cam_front = cv2.VideoCapture(0, cv2.CAP_DSHOW)  # front camera

for cam in [cam_back, cam_front]: # images taken at 1920x1080 pixels
    cam.set(cv2.CAP_PROP_FRAME_WIDTH, 1920)
    cam.set(cv2.CAP_PROP_FRAME_HEIGHT, 1080)

if not cam_back.isOpened() or not cam_front.isOpened():
    print("Error in opening camera(s)")
    exit()

def save_to_excel(serial):
    file_exists = os.path.exists(EXCEL_FILE)

    if not file_exists:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Serials"
        ws.append(["serial_number", "folder"])
        wb.save(EXCEL_FILE)

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # duplicate prevention
    for row in ws.iter_rows(values_only=True): # iterates through rows
        if str(row[0]) == serial: 
            print("Duplicate skipped: " + serial)
            return False

    ws.append([serial, serial])

    wb.save(EXCEL_FILE)
    print("Saved to Excel: " + serial)
    return True

# auto roi detection

def find_serial_roi(frame):
    # ocr works better on grayscale
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    # smooth noise
    blur = cv2.GaussianBlur(gray, (5,5), 0)
    # picks best treshold and creates contrast by making white text w black background
    _, bw = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    # small gaps between characters and creates box around serial area
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (25,25))
    closed = cv2.morphologyEx(bw, cv2.MORPH_CLOSE, kernel)
    # possible "object" in the binary image
    contours, _ = cv2.findContours(closed, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    if not contours:
        return None

    h, w = frame.shape[:2]
    candidates = [] # store valid ROI options

    for cnt in contours: 
        # x,y --> top left corner
        # cw, ch --> width and height of contour box
        x, y, cw, ch = cv2.boundingRect(cnt)
        # ignore small objects
        if cw < 120 or ch < 40:
            continue
        # ignore squares
        aspect = cw / float(ch)
        if 0.8 <= aspect <= 1.2:
            continue
        # prefer right side candidates
        center_x = x + cw/2
        right_bias = center_x > w * 0.45 # 45% of width 
        # store position, size, if it's on the right, and area
        candidates.append((x, y, cw, ch, right_bias, cw*ch))

    if not candidates:
        return None
    # sort candidates as right side, then larger area
    candidates.sort(key=lambda c: (c[4], c[5]))
    x, y, cw, ch, _, _ = candidates[-1] # best candidate after sorting
    return (x, y, cw, ch)

def extract_serial(text):
    nums = re.findall(r"\d+", text)
    for n in nums:
        if len(n) == 11: # must be 11 digits/characters
            return n

    joined = "".join(nums)
    if len(joined) == 11:
        return joined

    return None
# main loop
while True:
    # read frames
    ret_back, frame_back = cam_back.read()
    ret_front, frame_front = cam_front.read()

    if not ret_back or not ret_front:
        continue

    roi_box = find_serial_roi(frame_back)

    if roi_box:
        x, y, w, h = roi_box
        roi = frame_back[y:y+h, x:x+w] # crops part of image with the serial number

        # draw roi on the main view
        cv2.rectangle(frame_back, (x,y), (x+w,y+h), (0,255,0), 3)

        # convert bgr (from open cv) to rgb (for easyocr)
        rgb = cv2.cvtColor(roi, cv2.COLOR_BGR2RGB)
        # run ocr nn which returns a list of text regions, followed by confidence
        results = reader.readtext(rgb, detail=1)

        text = " ".join([t[1] for t in results])
        print("OCR:", text)

        serial = extract_serial(text)

        if serial:
            print("Detected serial:", serial)

            # create folder for this serial
            folder = os.path.join(BASE_DIR, serial)
            os.makedirs(folder, exist_ok=True)

            # save serial only once
            if save_to_excel(serial):
                # save back and front images
                cv2.imwrite(os.path.join(folder, serial+"_back.png"), frame_back)
                cv2.imwrite(os.path.join(folder, serial+"_front.png"), frame_front)
                print("Saved images for " + serial)

            time.sleep(1.0)

    display = cv2.resize(frame_back, (800, 800))
    cv2.imshow("Back Camera", display)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cam_back.release()
cam_front.release()
cv2.destroyAllWindows()
